from canvasapi import Canvas
from datetime import datetime
from openpyxl import load_workbook
import os

API_URL = "https://schulich.instructure.com/"
# Change this link based on the School you attend.

API_KEY = os.environ['CANVAS_API_KEY']

# Initialize a new Canvas object
canvas = Canvas(API_URL, API_KEY)

def retrieveAssignments(course):
    """ Retrieves all the assignments for a course, and related info. """
    course = canvas.get_course(course)
    assignments = course.get_assignments()
    courseAssignments = {}
    currentAssignment = {}
    for assignment in assignments:
        name = assignment.name
        dueDateRaw = assignment.due_at
        # Unpacking and formatting date.
        if dueDateRaw is None: # Usually if due date hasn't been set.
            dueDate = dueTime = "N/A"
        else:
            dueDateObj = datetime.strptime(dueDateRaw, '%Y-%m-%dT%H:%M:%SZ')
            dueDate = dueDateObj.strftime("%Y-%m-%d")
            dueTime = dueDateObj.strftime("%H:%M")

        weight = assignment.points_possible
        assignmentInfo = {
            "course":course.name[0:11],
            "dueDate": dueDate,
            "dueTime": dueTime,
            "weight": weight
        }
        currentAssignment[name] = assignmentInfo
        courseAssignments.update(currentAssignment)
    return courseAssignments

def writeToSheet():
    """ Writes all assignment information to a spreadsheet."""
    workbook = load_workbook(filename='Canvas Assignments (Master).xlsx')
    sheet = workbook.active
    allCourses = canvas.get_courses()
    courseIDS = []
    count = 1
    for items in allCourses:
        if items.id != 3967: # Schulich Academic Dishonesty Course
            courseIDS.append(items.id)
    for i in range(len(courseIDS)):
        assignments = retrieveAssignments(courseIDS[i])
        for assignment, info in assignments.items():
            count += 1
            sheet.cell(row=count, column=7, value=assignment)
            sheet.cell(row=count, column=4, value=info['course'])
            sheet.cell(row=count, column=1, value=info['dueDate'])
            sheet.cell(row=count, column=3, value=info['dueTime'])
            sheet.cell(row=count, column=11, value=info['weight'])
    workbook.save(filename="Canvas-Assignments.xlsx")
writeToSheet()
