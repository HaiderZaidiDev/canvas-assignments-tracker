from django.shortcuts import render
from .forms import AccessTokenForm
from canvasapi import Canvas
from datetime import datetime
from openpyxl import load_workbook
import os
import mimetypes
from django.http import HttpResponse
from pytz import timezone
import secrets

def retrieveAssignments(course, token):
    """ Retrieves all the assignments for a course, and related info. """
    API_URL = "https://schulich.instructure.com/"
    # Change this link based on the School you attend.
    try:
        canvas = Canvas(API_URL, token)
    except:
        handler500View()

    # Initialize a new Canvas object

    course = canvas.get_course(course)
    assignments = course.get_assignments()
    courseAssignments = {}
    currentAssignment = {}
    for assignment in assignments:
        name = assignment.name
        dueDateRaw = assignment.due_at
        # Unpacking and formatting date.
        if dueDateRaw is None: # Usually if due date hasn't been set.
            dueDate = dueTime = "N/A"
        else:
            dueDateObj = datetime.strptime(dueDateRaw, '%Y-%m-%dT%H:%M:%SZ')
            dueDateEST = dueDateObj.replace(tzinfo=timezone('UTC')).astimezone(timezone('EST5EDT'))
            dueDate = dueDateEST.strftime("%Y-%m-%d")
            dueTime = dueDateEST.strftime("%H:%M")

        weight = assignment.points_possible
        assignmentInfo = {
            "course":course.name[0:11],
            "dueDate": dueDate,
            "dueTime": dueTime,
            "weight": weight
        }
        currentAssignment[name] = assignmentInfo
        courseAssignments.update(currentAssignment)
    return courseAssignments

def writeToSheet(token):
    """ Writes all assignment information to a spreadsheet."""
    API_URL = "https://schulich.instructure.com/"
    # Change this link based on the School you attend.
    API_KEY = token
    try:
        canvas = Canvas(API_URL, token)
    except:
        return render(request, 'error500.html', {})

    workbook = load_workbook(filename='/var/www/canvas-assignments-to-excel/assignmentTracker/pages/sheets/master-sheets/Canvas Assignments (Master).xlsx')
    sheet = workbook.active
    allCourses = canvas.get_courses()
    courseIDS = []
    count = 1

    # Getting IDS of all courses the user is enrolled in
    for items in allCourses:
        if items.id != 3967: # Schulich Academic Dishonesty Course
            courseIDS.append(items.id)
    for i in range(len(courseIDS)):
        assignments = retrieveAssignments(courseIDS[i], token)
        for assignment, info in assignments.items():
            count += 1
            sheet.cell(row=count, column=7, value=assignment)
            sheet.cell(row=count, column=4, value=info['course'])
            sheet.cell(row=count, column=1, value=info['dueDate'])
            sheet.cell(row=count, column=3, value=info['dueTime'])
            sheet.cell(row=count, column=11, value=info['weight'])
    hex_token = secrets.token_hex(4)
    workbook.save(filename=f"/var/www/canvas-assignments-to-excel/assignmentTracker/pages/sheets/user-sheets/Canvas-Assignments ({hex_token}).xlsx")
    return hex_token
    #Slices characters for differentiation of filename.

def homeView(request):
    """ Renders home page, serves spreadsheet to user """
    form = AccessTokenForm(request.POST or None)
    if form.is_valid():
        form.save()
        tokenClean = form.cleaned_data['token']
        print(tokenClean)
        sheetName = f"Canvas-Assignments ({writeToSheet(tokenClean)}).xlsx"
        #writeToSheet(tokenClean)
        # Serving spreadsheet.
        with open(f"/var/www/canvas-assignments-to-excel/assignmentTracker/pages/sheets/user-sheets/{sheetName}",'rb') as spreadsheet:
            sheet = spreadsheet.read()
            response = HttpResponse(sheet)
            response['Content-Type'] = 'mimetype/submimetype'
            response['Content-Disposition'] = f'attachment; filename={sheetName}'
            return response
    return render(request, 'index.html', {'form':form})

def privacyView(request):
    return render(request, 'privacy.html', {})

def handler500View(request):
    """ Error 500 handling """
    print("500")
    return render(request, 'error500.html', {})

def handler404View(request, exception):
    """ Error 404 handling """
    print("404")
    return render(request, 'error404.html', {})
